import csv
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Style constants ───────────────────────────────────────────────────────────
HEADER_BG  = "1E1E2E"
HEADER_FG  = "FFFFFF"
ROW_ALT    = "F4F6FB"
ROW_NORMAL = "FFFFFF"
BORDER_CLR = "D0D7E3"

SCORE_STYLES = {
    "perfect":  ("1A7F4B", "D6F4E3"),
    "good":     ("7A5C00", "FFF3CD"),
    "possible": ("8A3A00", "FFE5CC"),
    "poor":     ("8B1A1A", "FFD6D6"),
}

def _border():
    s = Side(border_style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def _score_style(score):
    if score >= 80:   k = "perfect"
    elif score >= 60: k = "good"
    elif score >= 40: k = "possible"
    else:             k = "poor"
    return SCORE_STYLES[k]

def main():
    # ── Load CSV ──────────────────────────────────────────────────────────────
    try:
        with open("enriched_leads.csv", "r", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
    except FileNotFoundError:
        try:
            with open("businesses_data.csv", "r", encoding="utf-8") as f:
                rows = list(csv.DictReader(f))
                print("Notice: using businesses_data.csv as enriched_leads.csv was not found.")
        except FileNotFoundError:
            print("No CSV files found. Please run main.py first.")
            sys.exit(1)

    # ── Normalize scores + sort ───────────────────────────────────────────────
    for r in rows:
        try:
            r["fit_score"] = int(r.get("fit_score", 0))
        except ValueError:
            r["fit_score"] = 0
    rows.sort(key=lambda x: x["fit_score"], reverse=True)

    # ── Print top 20 ─────────────────────────────────────────────────────────
    print("\n--- Top 20 Companies ---")
    for i, row in enumerate(rows[:20]):
        print(f"{i+1}. {row.get('company_name', 'Unknown')} - Score: {row.get('fit_score', 0)}")

    if not rows:
        print("No data to export.")
        sys.exit(0)

    headers = list(rows[0].keys())

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Lead Results
    # ════════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Lead Results"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h.replace("_", " ").title())
        cell.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
    ws.row_dimensions[1].height = 28

    # Detect key column indices (flexible — works whatever order your CSV has)
    h_lower = [h.lower() for h in headers]
    score_col    = next((i+1 for i, h in enumerate(h_lower) if "fit_score" in h), None)
    reason_col   = next((i+1 for i, h in enumerate(h_lower) if "fit_reason" in h), None)
    desc_col     = next((i+1 for i, h in enumerate(h_lower) if "description" in h), None)

    wrap_cols = {c for c in [reason_col, desc_col] if c}

    emails_found = phones_found = total_score = 0

    for row_idx, row in enumerate(rows, start=2):
        bg = ROW_ALT if row_idx % 2 == 0 else ROW_NORMAL
        score = row.get("fit_score", 0)
        total_score += score
        if row.get("email"): emails_found += 1
        if row.get("phone"): phones_found += 1

        for col_idx, h in enumerate(headers, start=1):
            value = row.get(h, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _border()

            if col_idx == score_col and isinstance(score, int):
                fg, score_bg = _score_style(score)
                cell.fill      = PatternFill("solid", fgColor=score_bg)
                cell.font      = Font(name="Arial", bold=True, color=fg, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.font      = Font(name="Arial", size=9)
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True if col_idx in wrap_cols else False
                )

        ws.row_dimensions[row_idx].height = 55

    # Column widths — cap long ones, keep short ones tight
    COL_WIDTH_MAP = {
        "company_name": 28, "website": 26, "linkedin_url": 32,
        "industry": 22,     "description": 46, "company_si": 16,
        "location": 20,     "email": 26,   "phone": 16,
        "fit_score": 11,    "fit_reason": 44,
    }
    for col_idx, h in enumerate(headers, start=1):
        key = h.lower()[:10]
        width = COL_WIDTH_MAP.get(key) or COL_WIDTH_MAP.get(h.lower(), None)
        if width is None:
            # fallback: auto from content, capped at 40
            max_len = max((len(str(row.get(h, ""))) for row in rows), default=10)
            width = min(max(max_len + 2, len(h) + 2), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Excel table with built-in filter
    last_col = get_column_letter(len(headers))
    last_row = len(rows) + 1
    tab = Table(displayName="Leads", ref=f"A1:{last_col}{last_row}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False
    )
    ws.add_table(tab)

    # ════════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Score Summary
    # ════════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Score Summary")
    ws2.sheet_view.showGridLines = False

    summary_headers = ["Score Tier", "Count", "Companies"]
    tiers = [
        ("🟢 Perfect Fit  (80–100)", lambda s: s >= 80),
        ("🟡 Good Fit     (60–79)",  lambda s: 60 <= s < 80),
        ("🟠 Possible Fit (40–59)",  lambda s: 40 <= s < 60),
        ("🔴 Poor Fit     (0–39)",   lambda s: s < 40),
    ]

    for col_idx, h in enumerate(summary_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
    ws2.row_dimensions[1].height = 26
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 65

    for row_idx, (label, fn) in enumerate(tiers, start=2):
        matched = [r.get("company_name", "") for r in rows
                   if isinstance(r.get("fit_score"), int) and fn(r["fit_score"])]
        bg = ROW_ALT if row_idx % 2 == 0 else ROW_NORMAL
        for col_idx, val in enumerate([label, len(matched), ", ".join(matched)], start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = _border()
        ws2.row_dimensions[row_idx].height = 22

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save("final_leads.xlsx")
    print(f"\nSaved {len(rows)} companies to final_leads.xlsx")

    # ── Summary ───────────────────────────────────────────────────────────────
    avg_score = total_score / len(rows) if rows else 0
    print("\n--- Summary ---")
    print(f"Total companies : {len(rows)}")
    print(f"Average score   : {avg_score:.1f}")
    print(f"Emails found    : {emails_found}")
    print(f"Phones found    : {phones_found}")

if __name__ == "__main__":
    main()